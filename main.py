import openpyxl

if __name__ == '__main__':

    # 設定
    # 目次を生成するエクセルファイルの名前
    inputWorkBookName = 'sampleSheet.xlsx'
    # 目次を生成した後保存するファイル名
    outputWorkBookName = 'sampleSheetOutput.xlsx'

    # ワークブックを取得
    wb = openpyxl.load_workbook(inputWorkBookName)

    # シート名一覧をlistで取得
    sheetNames = wb.sheetnames

    # 0番目に「index」というシートを生成
    wb.create_sheet(index=0, title='index')

    # 「index」シートオブジェクトを取得
    indexSheet = wb['index']

    # シート名とハイパーリンクを設定
    for i, sheetName in enumerate(sheetNames):
        indexSheet.cell(row=i+1, column=1).value = sheetName
        hyperLinkUrl = outputWorkBookName + '#' + sheetName + '!A1'
        indexSheet.cell(row=i+1, column=1).hyperlink = hyperLinkUrl

    # シートを保存
    wb.save(outputWorkBookName)